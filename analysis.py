import os
import re
import argparse
import myers
import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from datetime import *
from pathlib import Path


EXTRACT_EXPR = r'\*+(?:BEGIN|END)\*+\n'
STATEPA_EXPR = r'(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(\w+)\ +(\w+)\ +(0x[0-9a-f]+) (\w+)'
INNER_TRACE_EXPR = r'(0x[0-9a-f]+)\ +(0x[0-9a-f]+)\ +(\d+)\ +(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(0x[0-9a-f]+)\ +(0x[0-9a-f]+)\ +(0x[0-9a-f]+)\ +(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(\d+)\ +(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}\.\d{3})\ +(\d+)\ +(\d+)\ +(\d+)\ +(\d+)\ +(0x[0-9a-f]+)\ +([0-9a-f]+)\ ([0-9a-f]+)\ +([0-9a-f]+)\ +([0-9a-f]+)\ +([0-9a-f]+)\ +([0-9a-f]+)\ +([0-9a-f]+)\ +([0-9a-f]+)'

# print([test_time(res[0][0], res[1][0]), test_time(res[1][0], res[2][0])])
def test_time(t1, t2):
    (h1, m1, s1, ms1) = re.compile(r'(\d{1,2}):(\d{1,2}):(\d{1,2}).(\d{3})').findall(t1)[0]
    (h2, m2, s2, ms2) = re.compile(r'(\d{1,2}):(\d{1,2}):(\d{1,2}).(\d{3})').findall(t2)[0]
    r1 = int(h1)*60*60*1000 + int(m1)*60*1000 + int(s1)*1000 + int(ms1)
    r2 = int(h2)*60*60*1000 + int(m2)*60*1000 + int(s2)*1000 + int(ms2)
    return r2 - r1

def test_extract(data):
    # data = ['2023-08-23 01:29:45.790 A B 0x00000000 C', '2023-08-23 01:29:45.810 A B 0x00000000 C',  '2023-08-23 01:30:04.010 A B 0x00000000 C']
    res = list(map(lambda x: re.compile(STATEPA_EXPR).findall(x)[0], data))
    return res

def test_extract_logs(file_path):
    with open(file_path, 'r') as f:
        return list(filter(lambda x:len(x) > 0, re.split(EXTRACT_EXPR, f.read())))

def extract_logs(file_reader):
    return list(filter(lambda x:len(x) > 0, re.split(EXTRACT_EXPR, file_reader.read())))

def statepa_style(worksheet, height, width):
    ft = Font(color="FF0000")

    for i in range(height):
        for index in [x + str(i + 1) for x in ['B','G','I']]:
            worksheet[index].font = ft
   
def inner_trace_style(worksheet, height, width):
    ft = Font(color="FF0000")

    for i in range(height):
        for index in [x + str(i + 1) for x in ['M','O']]:
            worksheet[index].font = ft

def inner_trace_diff_style(worksheet, height, width):
    ft = Font(color="FF0000")

    for i in range(height):
        for index in [x + str(i + 1) for x in ['I', 'L', 'O']]:
            worksheet[index].font = ft
    
    for i in range(1, height):
        if worksheet['F' + str(i + 1)].value == '+':
            for j in range(width):
                worksheet[chr(ord('A') + j) + str(i + 1)].fill = PatternFill(start_color='ffd7d5', end_color='ffd7d5', fill_type='solid')
        elif worksheet['F' + str(i + 1)].value == '-':
            for j in range(width):
                worksheet[chr(ord('A') + j) + str(i + 1)].fill = PatternFill(start_color='ccffd8', end_color='ccffd8', fill_type='solid')
        else:
            continue

def write_xlsx(rows, workbook, sheet_index, sheet_name, sheet_title, styles):
    workbook.create_sheet(sheet_name, sheet_index)
    worksheet = workbook[sheet_name]
    worksheet.append(sheet_title)

    for i in range(len(rows)):
        worksheet.append(rows[i])
    
    styles(worksheet, len(rows)+1, len(sheet_title))

    as_text = lambda x: str(x) if x is not None else ""
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (max(len(as_text(cell.value)) for cell in column_cells) + 2) * 0.95

def exract_record(log_content, words, regex):
    record = []
    for i in range(len(log_content)):
        if log_content[i].startswith(words):
            record.append(list(map(lambda x: re.compile(regex).findall(x)[0], filter(lambda x:len(x)>0, log_content[i].split('\n')[2:]))))
        else:
            continue
    return record

def merge_partial(data):
    base_time = data[0][0]
    res = []
    for item in data:
        cur_time = item[0]
        res.append([*item,][:1] + [time_differential(base_time, cur_time)] + [*item,][1:])      
        base_time = cur_time
    return res

def time_differential(base_time, cur_time):
    base_time = datetime.strptime(base_time, '%Y-%m-%d %H:%M:%S.%f')
    cur_time = datetime.strptime(cur_time, '%Y-%m-%d %H:%M:%S.%f')
    return str((cur_time - base_time) // timedelta(milliseconds=1))

def state_analysis(base_log, cur_log, out_path):
        base_record = exract_record(base_log, 'The cfg fsm information', STATEPA_EXPR)
        cur_record = exract_record(cur_log, 'The cfg fsm information', STATEPA_EXPR)

        wb = openpyxl.Workbook()

        for index, (base_row, cur_row) in enumerate(zip(base_record, cur_record)):
            # base_row = ['2023-08-23 01:29:45.790 A B 0x00000000 C', '2023-08-23 01:29:45.810 A B 0x00000000 C',  '2023-08-23 01:30:04.010 A B 0x00000000 C']
            # cur_row = ['2023-08-23 01:29:45.791 A B 0x00000000 C', '2023-08-23 01:29:45.812 A B 0x00000000 C',  '2023-08-23 01:30:04.013 A B 0x00000000 C']
            rows = list(map(lambda x: x[0] + [str(int(x[1][1]) - int(x[0][1]))] + x[1], zip(merge_partial(base_row), merge_partial(cur_row))))
            write_xlsx(rows, wb, index, "Record" + str(index), ['RunTime', 'BaseSegDiff', 'CurState', 'Event', 'ActionRslt', 'NewState', 'TimeDiff', 'RunTime', 'CurSegDiff', 'CurState', 'Event', 'ActionRslt', 'NewState'], statepa_style)
        
        wb.save(out_path)
        wb.close()
        print(f'File saved path: {out_path}')

def inner_trace_analysis(base_log, cur_log, out_path):
    base_record = exract_record(base_log, 'The inner config trace from cfg', INNER_TRACE_EXPR)[0]
    cur_record = exract_record(cur_log, 'The inner config trace from cfg', INNER_TRACE_EXPR)[0]
    # base_record = [('0x00000002', '0x00000002', '0x00000002', '0', '2023-08-25 01:29:05.653', '2023-08-25 01:29:05.653', '0x80cc001c', '0x00000000', '0x18150625', '2023-08-25 01:29:09.769', '782', '2023-08-25 01:29:05.653', '2023-08-25 01:29:09.758', '3965', '0', '0', '0', '0x00000000', 'ffffffff', '5f657361', '72636564', '69747079', '725f6e6f', '00007365', '00000000', '00000000'), ('0x00000002', '0x00000002', '0x00000002', '0', '2023-08-25 01:29:05.653', '2023-08-25 01:29:05.653', '0x80cc001c', '0x00000000', '0x18150625', '2023-08-25 01:29:09.769', '782', '2023-08-25 01:29:05.653', '2023-08-25 01:29:09.758', '3965', '0', '0', '0', '0x00000000', 'ffffffff', '00657361', '00000000', '00000000', '00000000', '00000000', '00000000', '00000000')]
    
    wb = openpyxl.Workbook()

    # [2, 7, 5, 6, 9, 14, 15, 16, 3, 4, 10, 11, -1, 8, -1]
    base_rows = list(map(lambda x: [x[2], x[7], x[5], x[6], x[9], x[14], x[15], x[16], x[3], x[4], x[10], x[11], time_differential(x[10], x[11]), x[8], time_differential(x[11], x[8])], base_record))
    cur_rows = list(map(lambda x: [x[2], x[7], x[5], x[6], x[9], x[14], x[15], x[16], x[3], x[4], x[10], x[11], time_differential(x[10], x[11]), x[8], time_differential(x[11], x[8])], cur_record))
    
    write_xlsx(base_rows, wb, 0, 'OldVersion', ['VrId', 'ClassId', 'SenderId', 'OpCode', 'DimRecCnt', 'UsrChg', 'SaveCkp', 'RetCode', 'InnerBeginTime', 'RequestTime', 'BeginDealTime', 'FinishEditTime', 'EditTimeDiff', 'RespondTime', 'CommitTimeDiff'], inner_trace_style)
    write_xlsx(cur_rows, wb, 1, 'NewVersion', ['VrId', 'ClassId', 'SenderId', 'OpCode', 'DimRecCnt', 'UsrChg', 'SaveCkp', 'RetCode', 'InnerBeginTime', 'RequestTime', 'BeginDealTime', 'FinishEditTime', 'EditTimeDiff', 'RespondTime', 'CommitTimeDiff'], inner_trace_style)

    base_class = [row[1] for row in base_rows]
    cur_class = [row[1] for row in cur_rows]
    diff = myers.diff(base_class, cur_class)
    
    i = 0; j = 0
    class_rows = []
    for index, item in enumerate(diff, 1):
        if item[0] == 'k':
            class_rows.append([index, base_rows[i][2], cur_rows[j][2], base_rows[i][1], cur_rows[j][1], '', base_rows[i][-3], cur_rows[j][-3], str(int(cur_rows[j][-3])-int(base_rows[i][-3])), base_rows[i][-1], cur_rows[j][-1], str(int(cur_rows[j][-1])-int(base_rows[i][-1])), base_rows[i][4], cur_rows[j][4], str(int(cur_rows[j][4])-int(base_rows[i][4]))])
            i += 1
            j += 1
        elif item[0] == 'i':
            class_rows.append([index, '', cur_rows[j][2], '', cur_rows[j][1], '+', '', cur_rows[j][-3], '', '', cur_rows[j][-1], '', '', cur_rows[j][4], ''])
            j += 1
        else:
            class_rows.append([index, base_rows[i][2], '', base_rows[i][1], '', '-', base_rows[i][-3], '', '', base_rows[i][-1], '', '', base_rows[i][4], '', ''])
            i += 1

    write_xlsx(class_rows, wb, 2, 'VersionDiff', ['ID', 'OldSenderId', 'NewSenderId', 'OldClassId', 'NewClassId', 'ClassFlag', 'OldEditTime', 'NewEditTime', 'EditTimeDiff', 'OldCommitTime', 'NewCommitTime', 'CommitTimeDiff', 'OldDimRecCnt', 'NewDimRecCnt', 'DimRecCntDiff'], inner_trace_diff_style)

    wb.save(out_path)
    wb.close()
    print(f'File saved path: {out_path}')

if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='vPerf v1.0.0')
    parser.add_argument('module')
    parser.add_argument('--base', '-b', type=argparse.FileType('r'), required=True, help='')
    parser.add_argument('--input', '-i', type=argparse.FileType('r'), required=True, help='')
    parser.add_argument('--output', '-o', help='')
    args = parser.parse_args()
    
    base_log = extract_logs(args.base)
    input_log = extract_logs(args.input)
    out_path = args.output

    # base_log = test_extract_logs("D:\\vPerf\data-1\\CFGLocatingLog\\CFGLocatingLog_0.txt")
    # input_log = test_extract_logs("D:\\vPerf\data-1\\CFGLocatingLog\\CFGLocatingLog_1.txt")

    # base_log = test_extract_logs("D:\\vPerf\\Display_Config_Diag_Inner-config_Trace_Pro_0.txt")
    # input_log = test_extract_logs("D:\\vPerf\\Display_Config_Diag_Inner-config_Trace_Pro_1.txt")
    # out_path = None

    if out_path == None:
        out_path = os.getcwd() + "\\" + args.module + '-' + datetime.now().strftime("%Y%m%d-%H%M%S")  + ".xlsx"
    
    if out_path[-5:] != '.xlsx':
        out_path += '.xlsx'
    
    if args.module == 'statepa':
        state_analysis(base_log, input_log, out_path)
    elif args.module == 'trace':
        inner_trace_analysis(base_log, input_log, out_path)
    else:
        print("Error: no module founded.")


# python analysis.py statepa -b ./data-1/CFGLocatingLog/CFGLocatingLog_0.txt -i ./data-1/CFGLocatingLog/CFGLocatingLog_1.txt
# python analysis.py statepa -b ./data-1/CFGLocatingLog/CFGLocatingLog_0.txt -i ./data-2/CFGLocatingLog/CFGLocatingLog_1.txt -o 1.xlsx

# python analysis.py trace -b ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_0.txt -i ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_1.txt
# python analysis.py trace -b ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_1.txt -i ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_0.txt
# python analysis.py trace -b ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_0.txt -i ./data-3/Display_Config_Diag_Inner-config_Trace_Pro_1.txt -o aaa.xlsx

# python analysis.py statepa -b fsm_662.txt -i fsm_290.txt